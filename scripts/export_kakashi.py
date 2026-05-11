"""
export_kakashi.py — Gera planilha XLSX simplificada (3 colunas) para o sistema Kakashi.

O Kakashi e um sistema externo Python que recebe esta planilha e gera PDFs
prontos para impressao a partir do mockup principal de cada anuncio.

A descricao gerada bate 1:1 com a descricao do produto pai no ERP
(build_erp_template.py), permitindo que o nome do PDF gerado pelo Kakashi
corresponda ao registro do produto no ERP - facilita auditoria visual e
busca rapida na pasta do Drive.

Uso:
    python scripts/export_kakashi.py <input.json> [--output-dir planilhas_geradas_shopee]

Entrada: mesmo JSON intermediario usado por build_shopee_template.py / build_erp_template.py.
Saida: <output-dir>/kakashi_<loja>_<data>.xlsx

Estrutura da planilha (copiada do template oficial):
    | Codigo (SKU)         | Descricao                          | URL imagem 1                    |
    | KIT3_JesusEOvelha    | Kit 3 Quadros - Jesus e Ovelha     | https://i.ibb.co/.../capa.jpg   |
    | Q1_OliveiraGetsemani | Quadro - Oliveira no Getsemani     | https://i.ibb.co/.../capa.jpg   |

1 linha por anuncio (nao por variacao). Linhas com imagem_capa invalida sao omitidas
e reportadas em avisos.
"""
import sys
import os
import json
import argparse
import shutil
from datetime import date

import openpyxl

from build_erp_template import TIPO_NOME, CONFIG


BASE_DIR = os.path.dirname(os.path.dirname(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, "planilhas_padrao", "kakashi_art_generator.xlsx")

EXTENSOES_VALIDAS = (".jpg", ".jpeg", ".png", ".webp")


def _link_valido(url: str) -> bool:
    if not url or not isinstance(url, str):
        return False
    if not (url.startswith("http://") or url.startswith("https://")):
        return False
    return url.lower().split("?")[0].endswith(EXTENSOES_VALIDAS)


def gerar_kakashi(input_json: dict, output_dir: str) -> tuple[str, list[str]]:
    """
    Gera o XLSX do Kakashi a partir do template oficial.
    Retorna (caminho_absoluto, avisos).
    """
    loja = input_json.get("loja", "LOJA")
    produtos = input_json.get("produtos", [])
    prefixo = CONFIG["lojas"].get(loja, {}).get("prefixo_descricao_erp", "")
    avisos: list[str] = []

    os.makedirs(output_dir, exist_ok=True)

    hoje = date.today().strftime("%Y-%m-%d")
    nome_arquivo = f"kakashi_{loja.lower()}_{hoje}.xlsx"
    caminho = os.path.join(output_dir, nome_arquivo)
    if os.path.exists(caminho):
        base = caminho[:-5]
        sufixo = 2
        while os.path.exists(f"{base}_{sufixo}.xlsx"):
            sufixo += 1
        caminho = f"{base}_{sufixo}.xlsx"

    shutil.copy2(TEMPLATE_PATH, caminho)
    wb = openpyxl.load_workbook(caminho)
    ws = wb["Planilha1"]

    linha = 2
    for produto in produtos:
        nome_sku = produto.get("nome_arte_sku", "")
        nome_display = produto.get("nome_arte_display", "")
        tipo = produto.get("tipo", "Q1")
        imagem_capa = produto.get("imagem_capa", "")

        if not nome_sku:
            avisos.append(f"Kakashi: produto sem nome_arte_sku ignorado ('{nome_display}')")
            continue

        if not _link_valido(imagem_capa):
            avisos.append(
                f"Kakashi: '{nome_display or nome_sku}' omitido — imagem_capa invalida ou ausente"
            )
            continue

        sku_base = f"{tipo}_{nome_sku}"
        tipo_nome = TIPO_NOME.get(tipo, tipo)
        desc_pai = f"{tipo_nome} - {prefixo}{nome_display}"
        ws.cell(row=linha, column=1, value=sku_base)
        ws.cell(row=linha, column=2, value=desc_pai)
        ws.cell(row=linha, column=3, value=imagem_capa)
        linha += 1

    wb.save(caminho)

    total = linha - 2
    print(f"[OK] Kakashi: {caminho} ({total} linhas)", file=sys.stderr)
    return caminho, avisos


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("input", help="Arquivo JSON com dados dos produtos")
    parser.add_argument("--output-dir", default="planilhas_geradas_shopee")
    args = parser.parse_args()

    with open(args.input, encoding="utf-8") as f:
        data = json.load(f)

    if os.path.isabs(args.output_dir):
        output_dir = args.output_dir
    else:
        output_dir = os.path.join(BASE_DIR, args.output_dir)

    caminho, avisos = gerar_kakashi(data, output_dir)
    for aviso in avisos:
        print(f"[AVISO] {aviso}", file=sys.stderr)
    print(caminho)
