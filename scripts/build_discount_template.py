"""
build_discount_template.py — Gera planilha de Promocao do Vendedor 25% off
a partir do arquivo `mass_update_sales_info.xlsx` exportado da Shopee Seller
Center (Meus Produtos -> Acoes -> Exportar).

Estrategia "regua de 25% off via tabela inflada" (validada 15/05/2026):
- Peter cadastra com preco INFLADO (col Preco da planilha de cadastro)
- Operador ativa Promocao do Vendedor 25% via este XLSX
- Shopee aplica CEILING(inflado x 0.75) = preco-alvo no card
- Cliente ve riscado + tag "-25%"; margem 20-25% preservada

Uso programatico:
    from scripts.build_discount_template import gerar_discount
    caminho, avisos = gerar_discount(input_xlsx_path, output_dir)

Uso CLI:
    python scripts/build_discount_template.py <input.xlsx> [--output-dir DIR]

Estrutura do XLSX gerado (9 colunas, formato template-discount oficial Shopee):
    1. ID do produto                       (espelhado do input)
    2. Nome do Produto (opcional)          (espelhado)
    3. Nº de Ref. Parent SKU (opcional)    (so na 1a variacao de cada pai)
    4. ID de variacao                      (espelhado)
    5. Variacao de nome (opcional)         (espelhado, ex: "40x60,Moldura Madeira")
    6. Nº de Ref. SKU (opcional)           (espelhado, ex: "Q1_4060MM_Salmo4610")
    7. Preco original (opcional)           (preco INFLADO da tabela canonica)
    8. Preco de desconto                   (preco-alvo)
    9. Limite de compra (opcional)         (vazio)
"""
import os
import sys
import re
import json
import argparse
from datetime import date

import openpyxl


BASE_DIR = os.path.dirname(os.path.dirname(__file__))

# Carrega config.json (precos inflados + precos_alvo)
CONFIG_PATH = os.path.join(BASE_DIR, "config.json")
with open(CONFIG_PATH, encoding="utf-8") as f:
    CONFIG = json.load(f)


# Headers exatos do template Shopee (validado contra discount xlsx gerado pelo
# Mercedes e contra o template-discount.xlsx oficial da Seller Center).
HEADERS_DISCOUNT = [
    "ID do produto",
    "Nome do Produto. (Opcional)",
    "Nº de Ref. Parent SKU. (Opcional)",
    "ID de variação",
    "Variação de nome. (Opcional)",
    "Nº de Ref. SKU. (Opcional)",
    "Preço original (opcional)",
    "Preço de desconto",
    "Limite de compra (Opcional)",
]

# SKU formato: {TIPO}_{TAM}{MOL}_{NOME} — ex: Q1_4060MM_Salmo4610
# Captura tipo, tamanho (4 dígitos: 3040 ou 4060) e codigo molde (2 letras).
_SKU_RE = re.compile(r'^(Q1|KIT2|KIT3)_(\d{4})([A-Z]{2})_')

# Mapeamento codigo molde -> tipo_mold da tabela de precos
_MOL_TO_TIPOMOLD = {
    "SM": "SM",
    "MP": "CM",
    "MB": "CM",
    "MM": "CM",
}


def _lookup_precos(sku: str) -> tuple[float, float] | None:
    """Extrai (inflado, alvo) de CONFIG a partir do SKU.

    Retorna None se o SKU nao casa com formato esperado, ou se tipo/tamanho/
    moldura nao estao na tabela canonica (Q1/KIT2/KIT3, 3040/4060, SM/MP/MB/MM).
    """
    if not sku:
        return None
    m = _SKU_RE.match(sku)
    if not m:
        return None
    tipo, tamanho, mol_sku = m.group(1), m.group(2), m.group(3)
    tipo_mold = _MOL_TO_TIPOMOLD.get(mol_sku)
    if not tipo_mold:
        return None
    try:
        inflado = CONFIG["precos"][tipo][tamanho][tipo_mold]
        alvo = CONFIG["precos_alvo"][tipo][tamanho][tipo_mold]
        return (inflado, alvo)
    except KeyError:
        return None


def gerar_discount(input_xlsx_path: str, output_dir: str) -> tuple[str, list[str]]:
    """Gera template-discount.xlsx a partir do mass_update_sales_info da Shopee.

    Args:
        input_xlsx_path: caminho do XLSX exportado pela Shopee Seller Center
            (Meus Produtos -> Acoes -> Exportar -> mass_update_sales_info).
        output_dir: pasta destino (sera criada se nao existir).

    Returns:
        (caminho_xlsx_gerado, lista_de_avisos)

    Avisos sao acumulados para SKUs cujo lookup falhou (formato errado,
    tipo nao suportado, tamanho/moldura fora da tabela canonica). A linha
    e mantida no output mas com Preco original/desconto vazios — operador
    decide se descarta ou ajusta manualmente.
    """
    if not os.path.isfile(input_xlsx_path):
        raise FileNotFoundError(f"Arquivo nao encontrado: {input_xlsx_path}")

    os.makedirs(output_dir, exist_ok=True)
    avisos: list[str] = []

    wb_in = openpyxl.load_workbook(input_xlsx_path, data_only=True)
    ws_in = wb_in.active
    if ws_in.max_row < 2:
        raise ValueError(
            "Planilha de input vazia ou sem dados (esperado header na linha 1 "
            "e variacoes a partir da linha 2)."
        )

    # Detectar quais colunas usar. O XLSX da Shopee tem MUITAS colunas
    # (mass_update_sales_info inclui preco, estoque, dimensoes, frete, etc.).
    # A gente le pelo HEADER os indices das 6 colunas que precisamos.
    header = [str(c.value).strip() if c.value else "" for c in ws_in[1]]

    def _find_col(candidates: list[str]) -> int:
        """Retorna indice (0-based) da coluna cujo header bate com qualquer
        candidato (case-insensitive, ignora pontuacao). Levanta se nao achar."""
        norm = lambda s: re.sub(r'[^a-z0-9]+', '', s.lower())
        normalized = [norm(h) for h in header]
        for cand in candidates:
            target = norm(cand)
            for i, h in enumerate(normalized):
                if target in h:
                    return i
        raise ValueError(
            f"Coluna nao encontrada no input. Candidatos: {candidates}. "
            f"Header detectado: {header}"
        )

    idx_id_produto = _find_col(["ID do produto", "ID Produto"])
    idx_nome       = _find_col(["Nome do Produto"])
    idx_parent_sku = _find_col(["Parent SKU"])
    idx_id_var     = _find_col(["ID de variação", "ID de variacao", "ID Variacao"])
    idx_var_nome   = _find_col(["Variação de nome", "Variacao de nome", "Variação"])
    idx_sku_ref    = _find_col(["Ref. SKU", "Ref SKU", "SKU"])

    # Cria XLSX novo do zero
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sheet"
    ws_out.append(HEADERS_DISCOUNT)

    total = 0
    sem_lookup = 0
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        id_produto = row[idx_id_produto]
        if not id_produto:
            continue  # linha vazia ou separadora

        nome       = row[idx_nome]      if idx_nome      < len(row) else ""
        parent_sku = row[idx_parent_sku] if idx_parent_sku < len(row) else None
        id_var     = row[idx_id_var]    if idx_id_var    < len(row) else ""
        var_nome   = row[idx_var_nome]  if idx_var_nome  < len(row) else ""
        sku_ref    = row[idx_sku_ref]   if idx_sku_ref   < len(row) else ""

        precos = _lookup_precos(str(sku_ref) if sku_ref else "")
        if precos is None:
            sem_lookup += 1
            inflado, alvo = None, None
            avisos.append(
                f"SKU '{sku_ref}' nao casou com tabela canonica "
                f"(produto {id_produto}, variacao {id_var}) — Preco original/desconto vazios."
            )
        else:
            inflado, alvo = precos

        ws_out.append([
            id_produto,
            nome,
            parent_sku,
            id_var,
            var_nome,
            sku_ref,
            inflado,
            alvo,
            None,  # Limite de compra (col 9, sempre vazio)
        ])
        total += 1

    # Salva
    hoje = date.today().strftime("%Y-%m-%d")
    nome_arquivo = f"discount_25off_{hoje}.xlsx"
    caminho = os.path.join(output_dir, nome_arquivo)
    if os.path.exists(caminho):
        base = caminho[:-5]
        sufixo = 2
        while os.path.exists(f"{base}_{sufixo}.xlsx"):
            sufixo += 1
        caminho = f"{base}_{sufixo}.xlsx"
    wb_out.save(caminho)

    print(
        f"[OK] Discount: {caminho} "
        f"({total} variacoes, {sem_lookup} sem lookup)",
        file=sys.stderr,
    )
    return caminho, avisos


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("input", help="Arquivo mass_update_sales_info.xlsx exportado da Shopee")
    parser.add_argument("--output-dir", default="planilhas_geradas_shopee")
    args = parser.parse_args()

    base = os.path.dirname(os.path.dirname(__file__))
    output_dir = args.output_dir
    if not os.path.isabs(output_dir):
        output_dir = os.path.join(base, output_dir)

    caminho, avisos = gerar_discount(args.input, output_dir)
    for aviso in avisos[:10]:
        print(f"[AVISO] {aviso}", file=sys.stderr)
    if len(avisos) > 10:
        print(f"[AVISO] ... e mais {len(avisos) - 10} avisos.", file=sys.stderr)
    print(caminho)
