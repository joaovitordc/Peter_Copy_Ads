"""
build_rejeitados_template.py — Gera planilha XLSX com produtos que foram
rejeitados durante o processamento do Peter (ex: capa nao subiu pro ImgBB).

Objetivo: separar produtos com problema dos produtos OK, pra que as planilhas
de output principais (Shopee/ERP/Kakashi) contenham APENAS dados validos. O
operador pode revisar a planilha de rejeitados pra entender o que falhou e
retentar manualmente esses produtos numa nova rodada.

Uso programatico:
    from scripts.build_rejeitados_template import gerar_rejeitados
    caminho = gerar_rejeitados(produtos_rejeitados, output_dir, loja)

Estrutura do XLSX (5 colunas):
    A: Nome da Arte (display)
    B: SKU base (sem prefixo tipo)
    C: Tipo (Q1, KIT2, KIT3)
    D: Motivo da rejeicao
    E: Loja
"""
import os
import sys
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def gerar_rejeitados(rejeitados: list[dict], output_dir: str, loja: str) -> str:
    """Gera planilha XLSX listando produtos rejeitados.

    Args:
        rejeitados: lista de dicts com chaves
            nome_arte_display, nome_arte_sku, tipo, motivo
        output_dir: pasta destino
        loja: identificador da loja (vai na coluna E)

    Returns:
        caminho absoluto do XLSX gerado
    """
    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rejeitados"

    # Header com estilo pra destacar
    headers = ["Nome da Arte", "SKU base", "Tipo", "Motivo", "Loja"]
    ws.append(headers)
    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill("solid", fgColor="DC2626")  # vermelho — sinal visual
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Larguras
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 12

    for r in rejeitados:
        ws.append([
            r.get("nome_arte_display", ""),
            r.get("nome_arte_sku", ""),
            r.get("tipo", ""),
            r.get("motivo", ""),
            loja,
        ])

    hoje = date.today().strftime("%Y-%m-%d")
    nome_arquivo = f"rejeitados_{loja.lower()}_{hoje}.xlsx"
    caminho = os.path.join(output_dir, nome_arquivo)
    if os.path.exists(caminho):
        base = caminho[:-5]
        sufixo = 2
        while os.path.exists(f"{base}_{sufixo}.xlsx"):
            sufixo += 1
        caminho = f"{base}_{sufixo}.xlsx"
    wb.save(caminho)

    print(f"[OK] Rejeitados: {caminho} ({len(rejeitados)} produtos)", file=sys.stderr)
    return caminho
