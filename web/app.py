"""
app.py — Servidor FastAPI para o frontend Etsy → Shopee.

Rotas:
  GET  /                               — Serve index.html
  GET  /api/lojas                      — Lista lojas disponíveis
  POST /api/processar                  — Inicia job (campo: arquivo, loja, modo)
  GET  /api/status/{job_id}            — Consulta status do job
  GET  /api/download/{job_id}/{tipo}   — Download do arquivo gerado (shopee|erp|kakashi)
  GET  /api/modelo/{tipo}              — Download do modelo de planilha de entrada

Executar:
  python -m uvicorn web.app:app --host 0.0.0.0 --port 8000 --reload
"""
import io
import os
import sys
import json
import uuid
import asyncio
import shutil
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional
from datetime import datetime, timedelta

from fastapi import FastAPI, File, Form, UploadFile, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv

# Carregar .env
BASE_DIR = Path(__file__).parent.parent
load_dotenv(BASE_DIR / ".env")

sys.path.insert(0, str(BASE_DIR))
sys.path.insert(0, str(BASE_DIR / "scripts"))  # pra importar build_discount_template etc direto

CONFIG_PATH = BASE_DIR / "config.json"
with open(CONFIG_PATH, encoding="utf-8") as f:
    CONFIG = json.load(f)

WEB_DIR = Path(__file__).parent
STATIC_DIR = WEB_DIR / "static"

IS_VERCEL = bool(os.environ.get("VERCEL"))
JOBS_DIR = Path("/tmp/jobs") if IS_VERCEL else WEB_DIR / "jobs"
JOBS_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="Etsy → Shopee")


# ── Gerenciamento de jobs ──────────────────────────────────────────────────

@dataclass
class JobState:
    job_id: str
    status: str = "aguardando"
    mensagem: str = "Aguardando início..."
    percent: int = 0
    shopee_path: Optional[str] = None
    erp_path: Optional[str] = None
    kakashi_path: Optional[str] = None
    rejeitados_path: Optional[str] = None
    produtos: int = 0
    rejeitados: int = 0
    avisos: list = field(default_factory=list)
    erro: Optional[str] = None
    criado_em: datetime = field(default_factory=datetime.now)
    # Pipeline split em 2 fases (revisao de capas obrigatoria):
    #   fase 1: parser + LLM + upload ImgBB com crop centralizado padrao
    #     -> status: "aguardando_confirmacao"
    #   operador revisa, ajusta crops (/api/recrop), marca descartes
    #   fase 2 (/api/confirmar): aplica descartes, gera 3 planilhas
    #     -> status: "concluido"
    loja: Optional[str] = None
    categoria: Optional[str] = None
    input_json: Optional[dict] = None
    produtos_rejeitados: list = field(default_factory=list)  # falha de upload na fase 1
    output_dir: Optional[str] = None
    descartados: list = field(default_factory=list)  # SKUs descartados


jobs: dict[str, JobState] = {}


def _state_path(job_id: str) -> Path:
    """Caminho do arquivo de estado persistido pra recuperar entre warm starts
    do mesmo container Lambda. NAO resolve multi-container — pra isso seria
    necessario Vercel Blob/KV (follow-up)."""
    return JOBS_DIR / job_id / "_state.json"


def _salvar_estado(job: JobState) -> None:
    """Persiste JobState como JSON. Idempotente, chamado a cada atualizacao.

    Em Vercel: /tmp/jobs/<id>/_state.json (sobrevive warm starts no mesmo container).
    Em local: web/jobs/<id>/_state.json.
    """
    try:
        path = _state_path(job.job_id)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump({
                "job_id":          job.job_id,
                "status":          job.status,
                "mensagem":        job.mensagem,
                "percent":         job.percent,
                "shopee_path":     job.shopee_path,
                "erp_path":        job.erp_path,
                "kakashi_path":    job.kakashi_path,
                "rejeitados_path": job.rejeitados_path,
                "produtos":        job.produtos,
                "rejeitados":      job.rejeitados,
                "avisos":          job.avisos,
                "erro":            job.erro,
                "criado_em":       job.criado_em.isoformat(),
                "loja":            job.loja,
                "categoria":       job.categoria,
                "input_json":      job.input_json,
                "produtos_rejeitados": job.produtos_rejeitados,
                "output_dir":      job.output_dir,
                "descartados":     job.descartados,
            }, f, ensure_ascii=False)
    except Exception as e:
        # Persistencia e best-effort — falha aqui nao pode quebrar o pipeline
        print(f"[AVISO] Falha ao persistir JobState {job.job_id}: {e}", file=sys.stderr)


def _carregar_estado(job_id: str) -> Optional[JobState]:
    """Recupera JobState do disco quando nao esta em memoria (cold start no
    mesmo container, ou simplesmente outra requisicao). Retorna None se nao
    existir ou estiver corrompido."""
    try:
        path = _state_path(job_id)
        if not path.exists():
            return None
        with open(path, encoding="utf-8") as f:
            d = json.load(f)
        return JobState(
            job_id=d["job_id"],
            status=d.get("status", "aguardando"),
            mensagem=d.get("mensagem", ""),
            percent=d.get("percent", 0),
            shopee_path=d.get("shopee_path"),
            erp_path=d.get("erp_path"),
            kakashi_path=d.get("kakashi_path"),
            rejeitados_path=d.get("rejeitados_path"),
            produtos=d.get("produtos", 0),
            rejeitados=d.get("rejeitados", 0),
            avisos=d.get("avisos", []),
            erro=d.get("erro"),
            criado_em=datetime.fromisoformat(d["criado_em"]) if d.get("criado_em") else datetime.now(),
            loja=d.get("loja"),
            categoria=d.get("categoria"),
            input_json=d.get("input_json"),
            produtos_rejeitados=d.get("produtos_rejeitados", []),
            output_dir=d.get("output_dir"),
            descartados=d.get("descartados", []),
        )
    except Exception as e:
        print(f"[AVISO] Falha ao carregar JobState {job_id}: {e}", file=sys.stderr)
        return None


def _limpar_jobs_antigos():
    limite = datetime.now() - timedelta(hours=2)
    para_remover = [jid for jid, j in jobs.items() if j.criado_em < limite]
    for jid in para_remover:
        job_dir = JOBS_DIR / jid
        if job_dir.exists():
            shutil.rmtree(job_dir, ignore_errors=True)
        del jobs[jid]


# ── Rotas ──────────────────────────────────────────────────────────────────

@app.get("/")
async def index():
    return FileResponse(STATIC_DIR / "index.html")


@app.get("/api/lojas")
async def get_lojas():
    """Retorna lista de lojas + categorias disponiveis por loja.

    Frontend usa o array `categorias` pra renderizar sub-selecao quando
    a loja tem mais de uma categoria (ex: AllQuadros = Padrao / Infantil).
    """
    lojas = []
    for nome, cfg in CONFIG["lojas"].items():
        categorias = [
            {"id": cat_id, "nome": cat.get("rotulo", cat_id.capitalize())}
            for cat_id, cat in cfg.get("categorias", {}).items()
        ]
        lojas.append({
            "id": nome,
            "nome": nome,
            "descricao": cfg.get("descricao", ""),
            "categoria_default": cfg.get("categoria_default", "padrao"),
            "categorias": categorias,
        })
    return {"lojas": lojas}


@app.post("/api/processar")
async def processar(
    background_tasks: BackgroundTasks,
    arquivo: UploadFile = File(...),
    loja: str = Form(...),
    modo: str = Form("links_com_imagens"),
    categoria: str = Form(""),
):
    """Inicia o processamento. modo: 'links' ou 'links_com_imagens'.
    categoria: 'padrao' | 'infantil' | ... (vazio = categoria_default da loja).
    """

    ext = Path(arquivo.filename).suffix.lower()
    if ext not in (".xlsx", ".xls", ".csv"):
        raise HTTPException(400, detail=f"Formato inválido '{ext}'. Use .xlsx, .xls ou .csv")

    if loja not in CONFIG["lojas"]:
        raise HTTPException(400, detail=f"Loja '{loja}' inválida.")

    if modo not in ("links", "links_com_imagens"):
        raise HTTPException(400, detail="Modo inválido. Use 'links' ou 'links_com_imagens'.")

    # Default = categoria_default da loja. Validacao concreta acontece em core.processar().
    loja_cfg = CONFIG["lojas"][loja]
    if not categoria:
        categoria = loja_cfg.get("categoria_default", "padrao")
    if categoria not in loja_cfg.get("categorias", {}):
        cats_validas = ", ".join(loja_cfg.get("categorias", {}).keys())
        raise HTTPException(400, detail=f"Categoria '{categoria}' inválida para '{loja}'. Use: {cats_validas}")

    job_id = str(uuid.uuid4())
    job_dir = JOBS_DIR / job_id
    job_dir.mkdir()

    input_path = job_dir / f"input{ext}"
    conteudo = await arquivo.read()
    with open(input_path, "wb") as f:
        f.write(conteudo)

    jobs[job_id] = JobState(
        job_id=job_id, loja=loja, categoria=categoria, output_dir=str(job_dir),
    )
    _salvar_estado(jobs[job_id])  # persistencia imediata pra polling pegar mesmo em outra request
    background_tasks.add_task(
        _executar_pipeline, job_id, str(input_path), loja, modo, str(job_dir), categoria,
    )
    _limpar_jobs_antigos()

    return {"job_id": job_id}


async def _executar_pipeline(
    job_id: str, filepath: str, loja: str, modo: str, output_dir: str, categoria: str = "",
):
    """Roda a FASE 1 do pipeline (parser → LLM → upload ImgBB crop centralizado).
    Termina marcando status `aguardando_confirmacao` — operador entao revisa
    capas, ajusta crops via /api/recrop, marca descartes e chama /api/confirmar
    pra rodar a fase 2 (gerar planilhas)."""
    from web.core import processar, ProcessamentoError

    job = jobs.get(job_id)
    if not job:
        return

    job.status = "processando"
    _salvar_estado(job)

    def progresso(msg: str, pct: int):
        if job_id in jobs:
            jobs[job_id].mensagem = msg
            jobs[job_id].percent = pct
            _salvar_estado(jobs[job_id])  # cada tick de progresso persiste

    try:
        resultado = await asyncio.to_thread(
            processar, filepath, loja, output_dir, modo, progresso, categoria,
        )
        n_ok = resultado["produtos"]
        n_rej = resultado.get("rejeitados", 0)
        job.input_json = resultado.get("input_json")
        job.produtos_rejeitados = resultado.get("produtos_rejeitados") or []
        job.produtos = n_ok
        job.rejeitados = n_rej
        job.avisos = resultado.get("avisos", [])
        job.status = "aguardando_confirmacao"
        if n_rej > 0:
            job.mensagem = f"{n_ok} produtos prontos pra revisao + {n_rej} rejeitados."
        else:
            job.mensagem = f"{n_ok} produtos prontos pra revisao."
        job.percent = 85

    except ProcessamentoError as e:
        job.status = "erro"
        job.erro = str(e)
        job.mensagem = str(e)
        job.avisos = getattr(e, "avisos", []) or []
    except Exception as e:
        job.status = "erro"
        job.erro = f"Erro inesperado: {e}"
        job.mensagem = "Ocorreu um erro inesperado. Tente novamente."
    finally:
        _salvar_estado(job)


async def _executar_fase2(job_id: str, descartes: list):
    """Roda a FASE 2 (gera as 3 planilhas). Disparada pelo /api/confirmar
    apos o operador revisar capas + ajustar crops + marcar descartes."""
    from web.core import gerar_planilhas, ProcessamentoError

    job = jobs.get(job_id)
    if not job:
        return
    if not job.input_json or not job.output_dir:
        return

    job.status = "gerando_planilhas"
    _salvar_estado(job)

    def progresso(msg: str, pct: int):
        if job_id in jobs:
            jobs[job_id].mensagem = msg
            jobs[job_id].percent = pct
            _salvar_estado(jobs[job_id])

    try:
        resultado = await asyncio.to_thread(
            gerar_planilhas,
            job.input_json,
            job.produtos_rejeitados or [],
            job.output_dir,
            descartes or [],
            list(job.avisos or []),
            progresso,
        )
        job.status = "concluido"
        n_ok = resultado["produtos"]
        n_rej = resultado.get("rejeitados", 0)
        if n_rej > 0:
            job.mensagem = f"{n_ok} produtos OK + {n_rej} rejeitados (veja a planilha 'Rejeitados')."
        else:
            job.mensagem = f"{n_ok} produtos processados com sucesso!"
        job.percent = 100
        job.shopee_path = resultado["shopee_path"]
        job.erp_path = resultado["erp_path"]
        job.kakashi_path = resultado.get("kakashi_path")
        job.rejeitados_path = resultado.get("rejeitados_path")
        job.produtos = n_ok
        job.rejeitados = n_rej
        job.avisos = resultado.get("avisos", [])
        # input_json final ja filtrado por descartes (post-fase2)
        job.input_json = resultado.get("input_json")
        job.descartados = list(set((job.descartados or []) + (descartes or [])))

    except ProcessamentoError as e:
        job.status = "erro"
        job.erro = str(e)
        job.mensagem = str(e)
        job.avisos = getattr(e, "avisos", []) or []
    except Exception as e:
        job.status = "erro"
        job.erro = f"Erro inesperado: {e}"
        job.mensagem = "Ocorreu um erro inesperado. Tente novamente."
    finally:
        _salvar_estado(job)


def _get_job(job_id: str) -> Optional[JobState]:
    """Retorna JobState do dict em memoria OU fallback do disco (/tmp/jobs/<id>/_state.json).
    Cobre o cenario onde o request HTTP cai em container Lambda warm que ja
    rodou o pipeline mas perdeu o dict (modulo recarregado). Cross-container
    100% ainda requer storage compartilhado (Vercel Blob/KV — follow-up)."""
    job = jobs.get(job_id)
    if job is None:
        job = _carregar_estado(job_id)
        if job is not None:
            jobs[job_id] = job  # cache de volta no dict pra reqs subsequentes
    return job


@app.get("/api/status/{job_id}")
async def status(job_id: str):
    job = _get_job(job_id)
    if not job:
        raise HTTPException(404, detail="Job não encontrado")
    return {
        "job_id":          job_id,
        "status":          job.status,
        "mensagem":        job.mensagem,
        "percent":         job.percent,
        "produtos":        job.produtos,
        "rejeitados":      job.rejeitados,
        "tem_rejeitados":  bool(job.rejeitados_path),
        "avisos":          job.avisos,
        "erro":            job.erro,
    }


@app.get("/api/download/{job_id}/{tipo}")
async def download(job_id: str, tipo: str):
    if tipo not in ("shopee", "erp", "kakashi", "rejeitados"):
        raise HTTPException(400, detail="Tipo deve ser 'shopee', 'erp', 'kakashi' ou 'rejeitados'")

    job = _get_job(job_id)
    if not job:
        raise HTTPException(404, detail="Job não encontrado")
    if job.status != "concluido":
        raise HTTPException(400, detail="Processamento ainda não concluído")

    caminho = {
        "shopee":     job.shopee_path,
        "erp":        job.erp_path,
        "kakashi":    job.kakashi_path,
        "rejeitados": job.rejeitados_path,
    }[tipo]
    if not caminho or not Path(caminho).exists():
        raise HTTPException(404, detail="Arquivo não encontrado")

    return FileResponse(
        caminho,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{Path(caminho).name}"'},
    )


@app.get("/api/produtos/{job_id}")
async def produtos(job_id: str):
    """Lista produtos processados pra revisao visual de capas + ajuste de crop.
    Aceita status `aguardando_confirmacao` (revisao pre-gerar planilhas) e
    `concluido` (revisao post-fato, retrocompat).
    """
    job = _get_job(job_id)
    if not job:
        raise HTTPException(404, detail="Job não encontrado")
    if job.status not in ("aguardando_confirmacao", "concluido"):
        raise HTTPException(400, detail="Processamento ainda não concluiu a fase 1")
    if not job.input_json:
        return {"produtos": [], "loja": job.loja, "suporte_descarte": False}

    lista = []
    for p in job.input_json.get("produtos", []):
        tipo = p.get("tipo", "Q1")
        sku_base = p.get("nome_arte_sku", "")
        lista.append({
            "sku_completo":          f"{tipo}_{sku_base}",
            "sku_base":              sku_base,
            "tipo":                  tipo,
            "display":               p.get("nome_arte_display", ""),
            "titulo":                p.get("titulo_shopee", ""),
            "imagem_capa":           p.get("imagem_capa", ""),
            "imagem_capa_original":  p.get("imagem_capa_original", ""),
            "descartado":            sku_base in (job.descartados or []),
        })
    return {
        "produtos":         lista,
        "loja":             job.loja,
        "categoria":        job.categoria,
        "status":           job.status,
        "suporte_descarte": True,
    }


@app.post("/api/confirmar")
async def confirmar(background_tasks: BackgroundTasks, payload: dict):
    """Dispara a fase 2 do pipeline: gera as 3 planilhas (Shopee/ERP/Kakashi)
    aplicando a lista de descartes. Se algum SKU foi descartado, tambem libera
    a loja correspondente no banco peter_skus_em_uso (libera o nome pra reuso).

    Body: { "job_id": str, "descartes": ["AnimaisOceano", ...] }  # descartes opcional
    """
    job_id = payload.get("job_id")
    descartes = payload.get("descartes") or []
    if not job_id:
        raise HTTPException(400, detail="job_id obrigatório")
    if not isinstance(descartes, list):
        raise HTTPException(400, detail="descartes deve ser lista")

    job = _get_job(job_id)
    if not job:
        raise HTTPException(404, detail="Job não encontrado")
    if job.status != "aguardando_confirmacao":
        raise HTTPException(400, detail=f"Job em status '{job.status}', esperava 'aguardando_confirmacao'")
    if not job.input_json:
        raise HTTPException(400, detail="Job sem input_json — refaça o processamento.")

    # Libera os SKUs descartados no banco antes de gerar as planilhas.
    # liberar_loja remove a loja do array lojas_cadastradas; se o array
    # ficar vazio, a linha eh deletada (libera o nome pra reuso futuro).
    if descartes:
        import sku_storage as _sku_storage
        for sku_base in descartes:
            try:
                _sku_storage.liberar_loja(sku_base, job.loja or "")
            except Exception as e:
                print(f"[AVISO] liberar_loja({sku_base}, {job.loja}) falhou: {e}", file=sys.stderr)

    # Dispara fase 2 em background — polling do frontend vai ver
    # "gerando_planilhas" -> "concluido"
    background_tasks.add_task(_executar_fase2, job_id, descartes)
    return {"job_id": job_id, "iniciado": True, "descartes": len(descartes)}


@app.post("/api/recrop")
async def recrop(payload: dict):
    """Reajusta o crop da capa de UM produto via Cropper.js no UI.

    Body: {
      job_id: str,
      sku_base: str,
      image_url: str,                  # URL original (pre-crop) da capa
      crop: {x: int, y: int, width: int, height: int}   # coordenadas em pixels da original
    }

    Baixa a imagem original, recorta com os params, sobe ImgBB e atualiza
    a URL da capa no input_json do job. Retorna { nova_url }.
    """
    job_id = payload.get("job_id")
    sku_base = payload.get("sku_base")
    image_url = payload.get("image_url") or ""
    crop = payload.get("crop") or {}

    if not job_id or not sku_base or not image_url:
        raise HTTPException(400, detail="job_id, sku_base e image_url obrigatórios")
    try:
        x = int(crop.get("x", 0))
        y = int(crop.get("y", 0))
        w = int(crop.get("width", 0))
        h = int(crop.get("height", 0))
    except (TypeError, ValueError):
        raise HTTPException(400, detail="crop.x/y/width/height devem ser numéricos")
    if w <= 0 or h <= 0:
        raise HTTPException(400, detail="crop.width e crop.height devem ser > 0")

    job = _get_job(job_id)
    if not job:
        raise HTTPException(404, detail="Job não encontrado")
    if job.status != "aguardando_confirmacao":
        raise HTTPException(400, detail=f"Job em status '{job.status}', esperava 'aguardando_confirmacao'")
    if not job.input_json:
        raise HTTPException(400, detail="Job sem input_json")

    # Roda upload em thread (download+crop+upload pode demorar ~3-5s)
    import upload_images as _upimg
    nova_url = await asyncio.to_thread(_upimg.upload_recrop, image_url, x, y, w, h)
    if not nova_url:
        raise HTTPException(502, detail="Falha no recrop (download/PIL/ImgBB). Veja logs do servidor.")

    # Atualiza a URL no produto correspondente
    atualizado = False
    for p in job.input_json.get("produtos", []):
        if p.get("nome_arte_sku") == sku_base:
            p["imagem_capa"] = nova_url
            atualizado = True
            break
    if not atualizado:
        raise HTTPException(404, detail=f"SKU '{sku_base}' não encontrado no job")

    _salvar_estado(job)
    return {"sku_base": sku_base, "nova_url": nova_url}


@app.get("/api/skus")
async def listar_skus(loja: str = "", q: str = "", sort: str = "criado_desc"):
    """Lista todos os SKUs cadastrados no banco (Supabase ou local).

    Query params:
      loja — filtra so SKUs cadastrados na loja informada (ex: PPJ).
      q    — busca substring (case-insensitive) em sku_base ou display.
      sort — ordenacao: criado_desc (default, mais recentes primeiro) |
             criado_asc | sku_asc | sku_desc.

    Retorna {backend, total, skus: [{sku_base, lojas, tipo, display, criado_em}]}
    """
    import sku_storage as _sku_storage
    try:
        skus_dict = _sku_storage.carregar()
    except Exception as e:
        raise HTTPException(500, detail=f"Erro ao carregar SKUs: {e}")

    q_lower = q.strip().lower()
    resultado = []
    for sku_base, info_sku in skus_dict.items():
        lojas = info_sku.get("lojas") or []
        if loja and loja not in lojas:
            continue
        display = info_sku.get("display", "") or ""
        if q_lower and q_lower not in sku_base.lower() and q_lower not in display.lower():
            continue
        resultado.append({
            "sku_base":  sku_base,
            "lojas":     lojas,
            "tipo":      info_sku.get("tipo", ""),
            "display":   display,
            "criado_em": info_sku.get("criado_em", ""),
        })

    # Ordena conforme sort param. criado_em vem como string ISO "YYYY-MM-DD",
    # entao sort lexicografico funciona pra datas. Tiebreaker = sku_base.
    if sort == "criado_asc":
        resultado.sort(key=lambda r: (r["criado_em"], r["sku_base"].lower()))
    elif sort == "sku_asc":
        resultado.sort(key=lambda r: r["sku_base"].lower())
    elif sort == "sku_desc":
        resultado.sort(key=lambda r: r["sku_base"].lower(), reverse=True)
    else:  # criado_desc (default)
        resultado.sort(key=lambda r: (r["criado_em"], r["sku_base"].lower()), reverse=True)

    return {
        "backend": _sku_storage.info(),
        "total":   len(resultado),
        "skus":    resultado,
        "sort":    sort,
    }


@app.delete("/api/skus/{sku_base}")
async def deletar_sku(sku_base: str):
    """Apaga uma linha INTEIRA da tabela peter_skus_em_uso — libera o nome
    pra reuso futuro independente de quantas lojas estavam no array.

    Operacao administrativa, irreversivel. Use /api/skus/{sku}/loja/{loja}
    se quiser remover so de uma loja especifica.
    """
    if not sku_base:
        raise HTTPException(400, detail="sku_base obrigatorio")
    import sku_storage as _sku_storage
    try:
        ok = _sku_storage.liberar(sku_base)
    except Exception as e:
        raise HTTPException(500, detail=f"Erro ao apagar SKU: {e}")
    if not ok:
        raise HTTPException(404, detail=f"SKU '{sku_base}' nao encontrado.")
    return {"sku_base": sku_base, "removido": True}


@app.delete("/api/skus/{sku_base}/loja/{loja}")
async def deletar_sku_loja(sku_base: str, loja: str):
    """Remove `loja` do array `lojas_cadastradas` do SKU. Se o array ficar
    vazio depois, a linha eh deletada automaticamente (libera o nome).
    """
    if not sku_base or not loja:
        raise HTTPException(400, detail="sku_base e loja obrigatorios")
    import sku_storage as _sku_storage
    try:
        ok = _sku_storage.liberar_loja(sku_base, loja)
    except Exception as e:
        raise HTTPException(500, detail=f"Erro ao remover loja: {e}")
    if not ok:
        raise HTTPException(404, detail=f"SKU '{sku_base}' nao tinha '{loja}' no array.")
    return {"sku_base": sku_base, "loja": loja, "removido": True}


# ─────────────────────────────────────────────────────────────────────────────
# Banco Kakashi — artes pra gerar PDFs sob demanda
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/kakashi")
async def listar_kakashi(loja: str = "", q: str = "", status: str = "todos",
                          sort: str = "criado_desc"):
    """Lista artes do banco peter_kakashi_artes com filtros.

    Query params:
      loja   — filtra por loja (PPJ, iPaper, AllQuadros) ou vazio = todas
      q      — busca substring em sku_base / sku_completo / descricao
      status — 'todos' (default) | 'pendente' | 'enviado'
      sort   — criado_desc (default) | criado_asc | sku_asc | sku_desc
    """
    if status not in ("todos", "pendente", "enviado"):
        raise HTTPException(400, detail="status invalido. Use: todos | pendente | enviado")
    if sort not in ("criado_desc", "criado_asc", "sku_asc", "sku_desc"):
        raise HTTPException(400, detail="sort invalido")

    import kakashi_storage as _kak
    try:
        artes = _kak.carregar(loja=loja, q=q, status=status, sort=sort)
    except Exception as e:
        raise HTTPException(500, detail=f"Erro ao carregar artes Kakashi: {e}")

    return {
        "backend": _kak.info(),
        "total":   len(artes),
        "artes":   artes,
        "sort":    sort,
        "status":  status,
    }


@app.post("/api/kakashi/baixar")
async def baixar_kakashi(payload: dict):
    """Gera planilha Kakashi seletiva com os SKUs informados e marca como
    enviado (auto-update enviado_kakashi_em = hoje).

    Body: { "sku_bases": ["AnimaisFofos", "Salmo23", ...] }

    Retorna XLSX como StreamingResponse pra download imediato.
    """
    sku_bases = payload.get("sku_bases") or []
    if not isinstance(sku_bases, list) or not sku_bases:
        raise HTTPException(400, detail="sku_bases deve ser lista nao vazia")

    import kakashi_storage as _kak
    # Carrega todas as artes pra filtrar pelas selecionadas (mais robusto que
    # 1 fetch por SKU — N requests vira 1)
    try:
        todas = _kak.carregar()  # sem filtros — carga completa
    except Exception as e:
        raise HTTPException(500, detail=f"Erro ao carregar artes: {e}")

    selecionadas_set = set(sku_bases)
    selecionadas = [a for a in todas if a.get("sku_base") in selecionadas_set]
    if not selecionadas:
        raise HTTPException(404, detail="Nenhum SKU encontrado pra esses sku_bases.")

    # Gera XLSX em memoria a partir do template oficial
    import io
    import openpyxl
    import shutil
    import tempfile
    from datetime import date as _date

    template_path = BASE_DIR / "planilhas_padrao" / "kakashi_art_generator.xlsx"
    if not template_path.exists():
        raise HTTPException(500, detail=f"Template Kakashi nao encontrado em {template_path}")

    # Como o template tem formatacao, faz copy + load + edit (mesma estrategia
    # do scripts/export_kakashi.py). Usa tempfile pra nao poluir o repo.
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        shutil.copy2(template_path, tmp.name)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb["Planilha1"]
        # Mantem a mesma ordem que veio do front (preserva ordem da selecao)
        ordem = {sku: i for i, sku in enumerate(sku_bases)}
        selecionadas.sort(key=lambda a: ordem.get(a.get("sku_base"), 9999))
        linha = 2
        for arte in selecionadas:
            ws.cell(row=linha, column=1, value=arte.get("sku_completo", ""))
            ws.cell(row=linha, column=2, value=arte.get("descricao", ""))
            ws.cell(row=linha, column=3, value=arte.get("imagem_capa", ""))
            linha += 1
        wb.save(tmp_path)

        with open(tmp_path, "rb") as f:
            conteudo = f.read()
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

    # Marca como enviado (so as que realmente entraram no XLSX)
    sku_bases_baixados = [a["sku_base"] for a in selecionadas]
    try:
        _kak.marcar_enviado(sku_bases_baixados)
    except Exception as e:
        print(f"[AVISO] Falha ao marcar enviados: {e}", file=sys.stderr)

    hoje = _date.today().strftime("%Y-%m-%d")
    nome_arquivo = f"kakashi_selecao_{hoje}.xlsx"
    return StreamingResponse(
        io.BytesIO(conteudo),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


@app.patch("/api/kakashi/{sku_base}")
async def atualizar_kakashi(sku_base: str, payload: dict):
    """Toggle do status enviado de uma arte.

    Body: { "enviado": true | false }
      true  -> marcar_enviado([sku_base]) — preenche enviado_kakashi_em = hoje
      false -> desmarcar(sku_base) — volta pra NULL (Pendente)
    """
    if not sku_base:
        raise HTTPException(400, detail="sku_base obrigatorio")
    enviado = payload.get("enviado")
    if not isinstance(enviado, bool):
        raise HTTPException(400, detail="enviado deve ser boolean")

    import kakashi_storage as _kak
    try:
        if enviado:
            n = _kak.marcar_enviado([sku_base])
            return {"sku_base": sku_base, "enviado": True, "atualizados": n}
        ok = _kak.desmarcar(sku_base)
        return {"sku_base": sku_base, "enviado": False, "atualizado": ok}
    except Exception as e:
        raise HTTPException(500, detail=f"Erro ao atualizar: {e}")


@app.delete("/api/kakashi/{sku_base}")
async def deletar_kakashi(sku_base: str):
    """Remove uma arte do banco Kakashi (linha deletada). Nao afeta o
    peter_skus_em_uso — sao bancos independentes."""
    if not sku_base:
        raise HTTPException(400, detail="sku_base obrigatorio")
    import kakashi_storage as _kak
    try:
        ok = _kak.liberar(sku_base)
    except Exception as e:
        raise HTTPException(500, detail=f"Erro ao apagar: {e}")
    if not ok:
        raise HTTPException(404, detail=f"Arte '{sku_base}' nao encontrada.")
    return {"sku_base": sku_base, "removido": True}


@app.get("/api/modelo/{tipo}")
async def modelo(tipo: str):
    """
    Gera e serve um modelo de planilha de entrada para download.
    tipo: 'links' | 'links_com_imagens'
    """
    if tipo not in ("links", "links_com_imagens"):
        raise HTTPException(400, detail="Tipo deve ser 'links' ou 'links_com_imagens'")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilha"

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(color="FFFFFF", bold=True)

    if tipo == "links":
        ws.title = "Links Etsy"
        headers = ["QUANTIDADE", "LINK DO ANÚNCIO"]
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 70
        # Linha de exemplo
        ws.append([3, "https://www.etsy.com/pt/listing/123456789/nome-do-produto"])
        example_row = ws[2]
        for cell in example_row:
            cell.font = Font(color="9CA3AF", italic=True)
        filename = "modelo_links_etsy.xlsx"
    else:
        ws.title = "Links + Imagens"
        headers = ["QUANTIDADE", "LINK DO ANÚNCIO", "IMAGEM CAPA", "IMAGEM 1", "IMAGEM 2", "IMAGEM 3"]
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 60
        for col in ["C", "D", "E", "F"]:
            ws.column_dimensions[col].width = 45
        ws.append([
            3,
            "https://www.etsy.com/pt/listing/123456789/nome-do-produto",
            "https://i.etsystatic.com/xxx/capa.jpg",
            "https://i.etsystatic.com/xxx/img1.jpg",
            "https://i.etsystatic.com/xxx/img2.jpg",
            "https://i.etsystatic.com/xxx/img3.jpg",
        ])
        example_row = ws[2]
        for cell in example_row:
            cell.font = Font(color="9CA3AF", italic=True)
        filename = "modelo_links_com_imagens.xlsx"

    # Inserir cabeçalho na linha 1 (deslocar exemplo para linha 2)
    ws.insert_rows(1)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Tooltip explicando os valores aceitos na coluna QUANTIDADE
    quantidade_comment = Comment(
        "Coluna QUANTIDADE:\n"
        "  1 = Solo (Q1)\n"
        "  2 = Kit 2 quadros (KIT2)\n"
        "  3 = Kit 3 quadros (KIT3)\n"
        "  Vazio = Detecção automática pelo LLM",
        "Peter",
    )
    quantidade_comment.width = 280
    quantidade_comment.height = 90
    ws["A1"].comment = quantidade_comment

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/desconto")
async def desconto(arquivo: UploadFile = File(...)):
    """
    Recebe um mass_update_sales_info.xlsx exportado da Shopee Seller Center
    e devolve o template-discount.xlsx preenchido com (Preço original, Preço
    de desconto) baseados na tabela canônica do config.json (Q1/KIT2/KIT3).

    Operação síncrona — sem job system. Retorno direto pra download.
    """
    ext = Path(arquivo.filename or "").suffix.lower()
    if ext not in (".xlsx",):
        raise HTTPException(400, detail=f"Formato inválido '{ext}'. Use .xlsx exportado pela Shopee.")

    # Salvar upload em /tmp pra ler com openpyxl
    tmp_dir = JOBS_DIR / f"desconto_{uuid.uuid4()}"
    tmp_dir.mkdir()
    input_path = tmp_dir / arquivo.filename
    conteudo = await arquivo.read()
    with open(input_path, "wb") as f:
        f.write(conteudo)

    try:
        import build_discount_template as _bd
        caminho, avisos = _bd.gerar_discount(str(input_path), str(tmp_dir))
    except Exception as e:
        # cleanup parcial e bubble up
        shutil.rmtree(tmp_dir, ignore_errors=True)
        raise HTTPException(400, detail=f"Erro ao gerar planilha de desconto: {e}")

    # Stream da resposta a partir de buffer (pra poder limpar o tmp_dir depois)
    with open(caminho, "rb") as f:
        buf = io.BytesIO(f.read())
    shutil.rmtree(tmp_dir, ignore_errors=True)
    buf.seek(0)

    # X-Avisos: warnings agregados (SKUs sem lookup) — frontend pode mostrar
    avisos_header = "|".join(avisos[:5])  # limitar tamanho do header
    if len(avisos) > 5:
        avisos_header += f"|... e mais {len(avisos) - 5}"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{Path(caminho).name}"',
            "X-Avisos": avisos_header[:2000],
            "X-Avisos-Count": str(len(avisos)),
            "Access-Control-Expose-Headers": "X-Avisos, X-Avisos-Count, Content-Disposition",
        },
    )


# ── Estáticos ──────────────────────────────────────────────────────────────
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
